# -*- coding: utf-8 -*-
"""
Carga REAL de produção: WEBCRM_PROD.xlsx -> Postgres (Supabase).

Adaptado de import_test_data.py (mesma limpeza, mesmo mapeamento aba->tabela), com
as diferenças abaixo:
  - Destino é Postgres (psycopg2), não SQLite -- schema/views/triggers já devem ter
    sido aplicados via backend/scripts/apply-schema.ts antes de rodar este script.
  - Floats "inteiros" do Excel (ex.: 1.0) são convertidos pra int -- Postgres não aceita
    "1.0" como literal implícito de coluna INTEGER (SQLite aceitava sem reclamar).
  - Duas correções de dados decididas com o usuário em 2026-08-12, aplicadas aqui:
    1) urls/carteira com cliente_id = 0 (sentinela de "sem cliente", 306 + 1 linhas na
       planilha real) -> migradas com cliente_id = NULL (schema.pg.sql já relaxou o
       NOT NULL dessas duas colunas pra isso).
    2) resp com cliente_id/pessoa_id apontando pra registro que não existe mais na
       planilha (7 linhas: clientes 196/50/615/623, pessoa 8) -> não migradas.
"""
import datetime
import os
import re

import openpyxl
import psycopg2
from psycopg2.extras import execute_values

XLSX_PATH = r"C:\Claude\WebCRM\WEBCRM_PROD.xlsx"

ERROR_TOKENS = {"#NAME?", "#REF!", "#VALUE!", "#N/A", "#DIV/0!", "#NULL!", "#NUM!", "#ERROR!"}

# Achado real na planilha de produção (2026-08-12): pelo menos 1 célula numérica (carteira,
# cart_emprestimos_mes) foi digitada como moeda formatada em texto ("R$ 1.631.651,25") em vez
# de número puro. Detecta esse padrão e converte pra float -- mais seguro que descartar o valor.
MONEY_RE = re.compile(r"^R\$\s*-?[\d.]+,\d{2}$")


def parse_moeda_br(v):
    neg = v.lstrip("R$").strip().startswith("-")
    digits = v.replace("R$", "").strip().lstrip("-").replace(".", "").replace(",", ".")
    num = float(digits)
    return -num if neg else num


def clean(value):
    if value is None:
        return None
    if isinstance(value, str):
        v = value.strip()
        if v == "" or v in ERROR_TOKENS:
            return None
        if MONEY_RE.match(v):
            return parse_moeda_br(v)
        return v
    if isinstance(value, datetime.datetime):
        if value.hour or value.minute or value.second:
            return value.strftime("%Y-%m-%d %H:%M:%S")
        return value.strftime("%Y-%m-%d")
    if isinstance(value, datetime.date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, datetime.time):
        return value.strftime("%H:%M:%S")
    # Excel/openpyxl devolve id/contagem como float (1.0) -- Postgres exige int de verdade
    # pra coluna INTEGER (SQLite era tolerante com isso, Postgres não).
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


# Mesmo mapeamento/ordem de import_test_data.py (respeita FK; urls entra logo após
# clientes/produtos/servidores pra disparar os triggers de cliente_status).
TABLES = [
    ("grupos_econ", "grupos_econ", [("grp_id", "grp_id"), ("grp_nome", "grp_nome")]),
    ("clientes", "clientes", [
        ("cliente_id", "cliente_id"), ("grp_id", "grp_id"), ("cliente_nome", "cliente_nome"),
        ("cliente_cnpj", "cliente_cnpj"), ("cliente_cnpj_fat", "cliente_cnpj_fat"),
        ("cliente_cnpj_number", "cliente_cnpj_number"), ("cliente_dat_bloqueio", "cliente_dat_bloqueio"),
        ("cliente_dia_venc_consumo", "cliente_dia_venc_consumo"),
        ("cliente_dia_venc_carteira", "cliente_dia_venc_carteira"),
        ("cliente_cod_github", "cliente_cod_github"), ("cliente_log", "cliente_log"),
    ]),
    ("produtos", "suites", [
        ("produto_id", "produto_id"), ("produto_area", "produto_area"), ("produto_nome", "produto_nome"),
        ("produto_detalhe", "produto_detalhe"), ("produto_suite", "produto_suite"),
        ("produto_tip_apuracao", "produto_tip_apuracao"), ("produto_sku", "produto_sku"),
        ("produto_franquia", "produto_franquia"), ("produto_grupo", "produto_grupo"),
        ("produto_preco", "produto_preco"), ("produto_recorrencia", "produto_recorrencia"),
        ("produto_regra_apuração", "produto_regra_apuracao"),
    ]),
    ("servidores", "server", [
        ("server_id", "server_id"), ("server_nome", "server_nome"), ("server_ambiente", "server_ambiente"),
        ("server_finalidade", "server_finalidade"), ("server_mysql", "server_mysql"),
        ("server_status", "server_status"), ("server_proc", "server_proc"),
        ("server_conteudo", "server_conteudo"), ("server_familia", "server_familia"),
    ]),
    ("list_resp_crono", "list_resp_crono", [("resp_id", "resp_id"), ("resp_nome", "resp_nome")]),
    ("list_tip_resp", "list_tip_resp", [("tip_resp", "tip_resp")]),
    ("list_url_status", "list_url_status", [("url_status", "url_status")]),
    ("indices_economicos", "index", [
        ("index_cod", "index_cod"), ("index_nome", "index_nome"), ("index_ano", "index_ano"),
        ("index_mes", "index_mes"), ("index_vlr", "index_vlr"),
    ]),
    ("fornecedores", "fornecedores", [
        ("fornecedor_id", "fornecedor_id"), ("fornecedor_area", "fornecedor_area"),
        ("fornecedor_nome", "fornecedor_nome"), ("fornecedor_cnpj", "fornecedor_cnpj"),
    ]),
    ("urls", "urls", [
        ("url_id", "url_id"), ("cliente_id", "cliente_id"), ("url_path", "url_path"),
        ("url_server", "server_id"), ("url_prod", "produto_id"),
        ("url_status", "url_status"), ("url_dt_status", "url_dt_status"), ("url_exc", "url_exc"),
        ("url_dt_exc", "url_dt_exc"), ("url_pasta_raiz", "url_pasta_raiz"),
        ("url_pasta_anexos", "url_pasta_anexos"), ("urb_bd", "urb_bd"), ("url_obs", "url_obs"),
    ]),
    ("pessoas", "pessoas", [
        ("pessoa_id", "pessoa_id"), ("pessoa_nome", "pessoa_nome"), ("pessoa_status", "pessoa_status"),
        ("pessoa_funcao", "pessoa_funcao"), ("pessoa_mail", "pessoa_mail"), ("pessoa_fone", "pessoa_fone"),
        ("pessoa_diretor", "pessoa_diretor"), ("pessoa_ger_exec", "pessoa_ger_exec"),
        ("pessoa_ger", "pessoa_ger"), ("pessoa_lider", "pessoa_lider"), ("pessoa_squad", "pessoa_squad"),
        ("pessoa_billable", "pessoa_billable"),
    ]),
    ("ferias_marcacao", "ferias_marcacao", [
        ("feriasm_id", "feriasm_id"), ("pessoa_id", "pessoa_id"), ("feriasm_tipo", "feriasm_tipo"),
        ("feriasm_ini", "feriasm_ini"), ("feriasm_fim", "feriasm_fim"),
    ]),
    ("contatos", "contatos", [
        ("contato_id", "contato_id"), ("cliente_id", "cliente_id"), ("contato_nome", "contato_nome"),
        ("contato_mail", "contato_mail"), ("contato_fone", "contato_fone"), ("contato_status", "contato_status"),
    ]),
    ("cart_mes", "cart_mes", [
        ("cart_mes_id", "cart_mes_id"), ("cart_ano_mes", "cart_ano_mes"), ("cart_vigencia_ativa", "cart_vigencia_ativa"),
    ]),
    ("precos_cliente", "precos_cliente", [
        ("pc_id", "pc_id"), ("cliente_id", "cliente_id"), ("produto_id", "produto_id"),
        ("cart_mes_id", "cart_mes_id"), ("pc_dat_niver", "pc_dat_niver"), ("pc_cod_index", "pc_cod_index"),
        ("pc_vlr_franquia", "pc_vlr_franquia"), ("pc_vlr_unit", "pc_vlr_unit"),
        ("pc_fx1_lim", "pc_fx1_lim"), ("pc_fx2_lim", "pc_fx2_lim"), ("pc_fx3_lim", "pc_fx3_lim"),
        ("pc_fx4_lim", "pc_fx4_lim"), ("pc_fx5_lim", "pc_fx5_lim"), ("pc_fx1_vlr", "pc_fx1_vlr"),
        ("pc_fx2_vlr", "pc_fx2_vlr"), ("pc_fx3_vlr", "pc_fx3_vlr"), ("pc_fx4_vlr", "pc_fx4_vlr"),
        ("pc_fx5_vlr", "pc_fx5_vlr"),
    ]),
    ("consumo_ana", "consumo_ana", [
        ("consumo_id", "consumo_id"), ("cliente_id", "cliente_id"), ("produto_id", "produto_id"),
        ("cart_mes_id", "cart_mes_id"), ("consumo_data", "consumo_data"), ("consumo_qtd", "consumo_qtd"),
        ("consumo_det", "consumo_det"), ("consumo_consit", "consumo_consit"),
    ]),
    ("faturamento", "faturamento", [
        ("fat_id", "fat_id"), ("cart_mes_id", "cart_mes_id"), ("cliente_id", "cliente_id"),
        ("fat_dat_venc", "fat_dat_venc"), ("fat_cod_venc_protheus", "fat_cod_venc_protheus"),
        ("fat_num_nfe", "fat_num_nfe"), ("fat_num_rps", "fat_num_rps"), ("fat_obs", "fat_obs"),
    ]),
    ("carteira", "carteira", [
        ("cart_id", "cart_id"), ("cliente_id", "cliente_id"), ("cart_mes_id", "cart_mes_id"),
        ("cart_qtd", "cart_qtd"), ("cart_vlr", "cart_vlr"), ("cart_pdd", "cart_pdd"),
        ("cart_sem_pdd", "cart_sem_pdd"), ("cart_fat", "cart_fat"), ("cart_qtd_mes", "cart_qtd_mes"),
        ("cart_emprestimos_mes", "cart_emprestimos_mes"), ("cart_ult_def", "cart_ult_def"),
        ("cart_data_base", "cart_data_base"), ("cart_dat_extracao", "cart_dat_extracao"),
        ("cart_rds", "cart_rds"), ("cart_db", "cart_db"), ("cart_prod", "cart_prod"),
        ("cart_url_plan_analitica", "cart_url_plan_analitica"),
    ]),
    ("resp", "resp", [
        ("resp_id", "resp_id"), ("cliente_id", "cliente_id"), ("resp_tipo", "resp_tipo"), ("pessoa_id", "pessoa_id"),
    ]),
    ("escala", "escala", [
        ("escala_id", "escala_id"), ("pessoa_id", "pessoa_id"), ("escala_data", "escala_data"),
        ("escala_hora_ini", "escala_hora_ini"), ("escala_hora_fim", "escala_hora_fim"),
    ]),
    ("portfolios", "port", [
        ("port_id", "port_id"), ("cliente_id", "cliente_id"), ("port_tipo", "port_tipo"),
        ("port_nome", "port_nome"), ("port_pm", "port_pm"), ("port_diretorio", "port_diretorio"),
        ("port_status", "port_status"), ("port_pdf", "port_pdf"),
    ]),
    ("crono", "crono", [
        ("crono_id", "crono_id"), ("port_id", "port_id"), ("crono_grupo", "crono_grupo"),
        ("crono_topico", "crono_topico"), ("crono_tipo", "crono_tipo"), ("crono_atividade", "crono_atividade"),
        ("crono_inicio", "crono_inicio"), ("crono_fim", "crono_fim"), ("crono_replan", "crono_replan"),
        ("crono_esforço", "crono_esforco"), ("crono_perc_atual", "crono_perc_atual"),
        ("crono_hh_orc", "crono_hh_orc"), ("crono_hh_real", "crono_hh_real"), ("crono_status", "crono_status"),
        ("resp_id", "resp_id"), ("crono_demanda_1", "crono_demanda_1"), ("crono_demanda_2", "crono_demanda_2"),
        ("crono_demanda_3", "crono_demanda_3"), ("crono_relat", "crono_relat"),
    ]),
    ("propostas", "propostas", [
        ("proposta_id", "proposta_id"), ("cliente_id", "cliente_id"), ("proposta_chamado", "proposta_chamado"),
        ("proposta_demanda", "proposta_demanda"), ("proposta_nome", "proposta_nome"),
        ("proposta_desc", "proposta_desc"), ("proposta_hh", "proposta_hh"), ("proposta_vlr", "proposta_vlr"),
        ("proposta_status", "proposta_status"), ("proposta_anexo", "proposta_anexo"),
    ]),
    ("forn_contratos", "forn_contratos", [
        ("forn_cont_id", "forn_cont_id"), ("fornecedor_id", "fornecedor_id"),
        ("forn_cont_num_contrato", "forn_cont_num_contrato"), ("forn_cont_tipo", "forn_cont_tipo"),
        ("forn_cont_nivel", "forn_cont_nivel"), ("forn_cont_aloc", "forn_cont_aloc"),
        ("forn_cont_qtd_prf", "forn_cont_qtd_prf"), ("forn_cont_desc", "forn_cont_desc"),
        ("forn_cont_tip_vlr", "forn_cont_tip_vlr"), ("forn_cont_vlr_mes", "forn_cont_vlr_mes"),
        ("forn_cont_dt_ini", "forn_cont_dt_ini"), ("forn_cont_dt_fim", "forn_cont_dt_fim"),
        ("forn_cont_ind_reaj", "forn_cont_ind_reaj"), ("forn_cont_status", "forn_cont_status"),
        ("pessoa_id", "pessoa_id"),
    ]),
    ("forn_pagadoria", "forn_pagadoria", [
        ("forn_pag_id", "forn_pag_id"), ("fornecedor_id", "fornecedor_id"), ("forn_pag_resp", "forn_pag_resp"),
        ("forn_pag_tipo", "forn_pag_tipo"), ("forn_pag_tipo_detalhado", "forn_pag_tipo_detalhado"),
        ("forn_pag_competencia", "forn_pag_competencia"), ("forn_pag_dat", "forn_pag_dat"),
        ("forn_pag_nome_prf", "forn_pag_nome_prf"), ("forn_pag_qtd", "forn_pag_qtd"),
        ("forn_pag_vlr_unit", "forn_pag_vlr_unit"), ("forn_pag_tot_bruto", "forn_pag_tot_bruto"),
        ("forn_pag_tot_liq", "forn_pag_tot_liq"),
        ("forn_pag_vlr_pag_cliente_bruto", "forn_pag_vlr_pag_cliente_bruto"),
        ("forn_pag_vlr_pag_cliente_liq", "forn_pag_vlr_pag_cliente_liq"),
        ("forn_pag_vlr_receita_bruta", "forn_pag_vlr_receita_bruta"),
        ("forn_pag_vlr_receita_liq", "forn_pag_vlr_receita_liq"), ("forn_pag_obs", "forn_pag_obs"),
    ]),
    ("anexos", "anexos", [
        ("anexo_id", "anexo_id"), ("cliente_id", "cliente_id"), ("fornecedor_id", "fornecedor_id"),
        ("anexo_nome", "anexo_nome"), ("anexo_data", "anexo_data"), ("anexo_arquivo", "anexo_arquivo"),
    ]),
]

# tabela -> colunas de destino que recebem NULL quando o valor de origem é 0 (sentinela de
# "sem vínculo" herdado da planilha, achado real na migração de 2026-08-12 -- ver docstring).
ZERO_AS_NULL = {
    "urls": {"cliente_id"},
    "carteira": {"cliente_id"},
}


def not_null_columns(conn, table_name):
    with conn.cursor() as cur:
        cur.execute(
            """SELECT column_name FROM information_schema.columns
               WHERE table_schema = 'public' AND table_name = %s AND is_nullable = 'NO'""",
            (table_name,),
        )
        return {r[0] for r in cur.fetchall()}


def load_usuarios(wb, conn):
    ws = wb["usuarios"]
    header = [c.value for c in ws[1]]
    idx = {name: i for i, name in enumerate(header)}

    usuarios_rows = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        user_mail = clean(row[idx["user_mail"]])
        if user_mail is None and all(v is None for v in row):
            continue
        user_nome = clean(row[idx["user_nome"]])
        user_status = clean(row[idx["user_status"]])
        if user_mail not in usuarios_rows:
            usuarios_rows[user_mail] = (user_nome, user_mail, user_status)

    with conn.cursor() as cur:
        for user_nome, user_mail, user_status in usuarios_rows.values():
            cur.execute(
                "INSERT INTO usuarios (user_nome, user_mail, user_status) VALUES (%s, %s, %s)",
                (user_nome, user_mail, user_status),
            )
    conn.commit()
    print(f"  {'usuarios':25s} <- {'usuarios':20s} {len(usuarios_rows):>6d} linhas")
    return len(usuarios_rows)


def backfill_cliente_tip_vlr(wb, conn):
    ws = wb["precos_cliente"]
    header = [c.value for c in ws[1]]
    idx = {name: i for i, name in enumerate(header)}
    if "pc_tip_vlr" not in idx or "cliente_id" not in idx:
        return

    tip_vlr_by_cliente = {}
    conflitos = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        cliente_id = clean(row[idx["cliente_id"]])
        tip_vlr = clean(row[idx["pc_tip_vlr"]])
        if cliente_id is None or tip_vlr is None:
            continue
        existente = tip_vlr_by_cliente.get(cliente_id)
        if existente is None:
            tip_vlr_by_cliente[cliente_id] = tip_vlr
        elif existente != tip_vlr:
            conflitos.add(cliente_id)

    with conn.cursor() as cur:
        for cliente_id, tip_vlr in tip_vlr_by_cliente.items():
            cur.execute("UPDATE clientes SET cliente_tip_vlr = %s WHERE cliente_id = %s", (tip_vlr, cliente_id))
    conn.commit()
    print(f"  {'clientes.cliente_tip_vlr':25s} <- {'precos_cliente':20s} {len(tip_vlr_by_cliente):>6d} clientes")
    if conflitos:
        print(f"  [AVISO] {len(conflitos)} cliente(s) com pc_tip_vlr divergente entre linhas de preço "
              f"(ficou o primeiro valor encontrado): {sorted(conflitos)}")


def load_valid_ids(wb, sheet, col):
    ws = wb[sheet]
    header = [c.value for c in ws[1]]
    idx = header.index(col)
    ids = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        v = clean(row[idx])
        if v is not None:
            ids.add(v)
    return ids


def main():
    conn = psycopg2.connect(os.environ["DATABASE_URL"])
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True, read_only=True)

    # ids validos, lidos direto da planilha (usados so pro filtro de `resp` orfao --
    # decisao tomada com o usuario: nao migrar linhas de resp cujo cliente/pessoa nao existe).
    valid_cliente_ids = load_valid_ids(wb, "clientes", "cliente_id")
    valid_pessoa_ids = load_valid_ids(wb, "pessoas", "pessoa_id")

    total_rows = 0
    for table_name, sheet_name, col_map in TABLES:
        ws = wb[sheet_name]
        header = [c.value for c in ws[1]]
        header_idx = {name: i for i, name in enumerate(header)}

        missing = [src for src, _ in col_map if src not in header_idx]
        if missing:
            print(f"  [AVISO] {sheet_name}: colunas não encontradas no cabeçalho: {missing}")

        usable_map = [(src, dst) for src, dst in col_map if src in header_idx]
        src_indexes = [header_idx[src] for src, _ in usable_map]
        dst_cols = [dst for _, dst in usable_map]

        nn_cols = not_null_columns(conn, table_name)
        not_null_positions = [i for i, dst in enumerate(dst_cols) if dst in nn_cols]
        zero_as_null_positions = [i for i, dst in enumerate(dst_cols) if dst in ZERO_AS_NULL.get(table_name, set())]
        cliente_id_pos = dst_cols.index("cliente_id") if "cliente_id" in dst_cols else None
        pessoa_id_pos = dst_cols.index("pessoa_id") if "pessoa_id" in dst_cols else None

        col_list = ", ".join(dst_cols)
        insert_sql = f"INSERT INTO {table_name} ({col_list}) VALUES %s"

        rows_to_insert = []
        skipped_notnull = 0
        skipped_orfao = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(row[i] is None for i in src_indexes):
                continue
            values = [clean(row[i]) for i in src_indexes]

            for i in zero_as_null_positions:
                if values[i] == 0:
                    values[i] = None

            # resp: pula linha cujo cliente_id/pessoa_id nao existe mais na planilha
            # (referencia genuinamente quebrada, nao sentinela -- decisao com o usuario).
            if table_name == "resp":
                if cliente_id_pos is not None and values[cliente_id_pos] is not None and values[cliente_id_pos] not in valid_cliente_ids:
                    skipped_orfao += 1
                    continue
                if pessoa_id_pos is not None and values[pessoa_id_pos] is not None and values[pessoa_id_pos] not in valid_pessoa_ids:
                    skipped_orfao += 1
                    continue

            if any(values[i] is None for i in not_null_positions):
                skipped_notnull += 1
                continue
            rows_to_insert.append(tuple(values))

        with conn.cursor() as cur:
            if rows_to_insert:
                execute_values(cur, insert_sql, rows_to_insert, page_size=1000)
        conn.commit()
        print(f"  {table_name:25s} <- {sheet_name:20s} {len(rows_to_insert):>6d} linhas")
        if skipped_notnull:
            print(f"  [AVISO] {table_name}: {skipped_notnull} linha(s) ignorada(s) (coluna NOT NULL vazia após limpeza)")
        if skipped_orfao:
            print(f"  [AVISO] {table_name}: {skipped_orfao} linha(s) ignorada(s) (cliente_id/pessoa_id não existe mais)")
        total_rows += len(rows_to_insert)

        # avança a sequence do IDENTITY (senão o próximo INSERT real via app colide com ids migrados)
        with conn.cursor() as cur:
            cur.execute(
                """SELECT kcu.column_name FROM information_schema.table_constraints tc
                   JOIN information_schema.key_column_usage kcu
                     ON kcu.constraint_name = tc.constraint_name AND kcu.table_schema = tc.table_schema
                   WHERE tc.table_schema='public' AND tc.table_name=%s AND tc.constraint_type='PRIMARY KEY'""",
                (table_name,),
            )
            pk_rows = cur.fetchall()
            if len(pk_rows) == 1 and pk_rows[0][0] in dst_cols:
                pk_col = pk_rows[0][0]
                cur.execute("SELECT pg_get_serial_sequence(%s, %s)", (table_name, pk_col))
                has_sequence = cur.fetchone()[0] is not None
                if has_sequence:
                    cur.execute(f"SELECT COALESCE(MAX({pk_col}), 0) FROM {table_name}")
                    max_val = cur.fetchone()[0]
                    cur.execute("SELECT setval(pg_get_serial_sequence(%s, %s), %s, %s)", (table_name, pk_col, max_val, max_val > 0))
        conn.commit()

        if table_name == "list_url_status":
            total_rows += load_usuarios(wb, conn)

        if table_name == "precos_cliente":
            backfill_cliente_tip_vlr(wb, conn)

    conn.close()
    print(f"\nTotal importado: {total_rows} linhas em {len(TABLES)} tabelas")


if __name__ == "__main__":
    main()
