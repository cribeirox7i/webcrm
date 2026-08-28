# -*- coding: utf-8 -*-
"""
Carga de MASSA DE TESTE a partir de WEBCRM.xlsx pro SQLite do backend.

Isso NAO e o script de migracao real (a planilha continua em producao e
sendo alimentada -- a migracao definitiva vai precisar de limpeza mais
cuidadosa quando chegar a hora do corte). Aqui o objetivo e só povoar o
banco local com volume realista pra testar o backend.

Limpeza minima aplicada:
  - strings de erro do Sheets/Excel (#NAME?, #REF!, etc) -> NULL
  - string vazia -> NULL
  - datas/horas (objetos datetime/date/time do openpyxl) -> texto ISO-8601
  - so as colunas mantidas no schema.sql sao carregadas; formulas removidas
    (JOIN, agregacao, regra de negocio) ficam de fora, pois viram VIEW/trigger/backend
"""
import datetime
import os
import re
import sqlite3

import openpyxl

XLSX_PATH = r"C:\Users\carlo\Downloads\WEBCRM.xlsx"
DB_PATH = r"C:\Claude\WEBCRM\backend\data\webcrm.sqlite"
SQL_DIR = r"C:\Claude\WEBCRM"

ERROR_TOKENS = {"#NAME?", "#REF!", "#VALUE!", "#N/A", "#DIV/0!", "#NULL!", "#NUM!", "#ERROR!"}


def clean(value):
    if value is None:
        return None
    if isinstance(value, str):
        v = value.strip()
        if v == "" or v in ERROR_TOKENS:
            return None
        return v
    if isinstance(value, datetime.datetime):
        if value.hour or value.minute or value.second:
            return value.strftime("%Y-%m-%d %H:%M:%S")
        return value.strftime("%Y-%m-%d")
    if isinstance(value, datetime.date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, datetime.time):
        return value.strftime("%H:%M:%S")
    return value


# cada tabela: (aba de origem, [(cabecalho_na_planilha, coluna_no_banco), ...])
# a ordem da lista abaixo tambem e a ordem de carga (respeita FKs; urls entra
# antes de nada que dependa de cliente_status pra deixar os TRIGGERS calcularem).
TABLES = [
    ("grupos_econ", "grupos_econ", [
        ("grp_id", "grp_id"), ("grp_nome", "grp_nome"),
    ]),
    ("clientes", "clientes", [
        ("cliente_id", "cliente_id"), ("grp_id", "grp_id"), ("cliente_nome", "cliente_nome"),
        ("cliente_cnpj", "cliente_cnpj"), ("cliente_cnpj_fat", "cliente_cnpj_fat"),
        ("cliente_cnpj_number", "cliente_cnpj_number"), ("cliente_dat_bloqueio", "cliente_dat_bloqueio"),
        ("cliente_dia_venc_consumo", "cliente_dia_venc_consumo"),
        ("cliente_dia_venc_carteira", "cliente_dia_venc_carteira"),
        ("cliente_cod_github", "cliente_cod_github"), ("cliente_log", "cliente_log"),
        # cliente_tip_vlr não vem da aba "clientes" -- backfill_cliente_tip_vlr() abaixo
        # preenche a partir da coluna pc_tip_vlr da aba "precos_cliente" após a carga
    ]),
    ("produtos", "suites", [
        ("produto_id", "produto_id"), ("produto_area", "produto_area"), ("produto_nome", "produto_nome"),
        ("produto_detalhe", "produto_detalhe"), ("produto_suite", "produto_suite"),
        ("produto_tip_apuracao", "produto_tip_apuracao"), ("produto_sku", "produto_sku"),
        ("produto_franquia", "produto_franquia"), ("produto_grupo", "produto_grupo"),
        ("produto_preco", "produto_preco"), ("produto_recorrencia", "produto_recorrencia"),
        ("produto_regra_apuração", "produto_regra_apuracao"),
    ]),
    ("servidores", "server", [
        ("server_id", "server_id"), ("server_nome", "server_nome"), ("server_ambiente", "server_ambiente"),
        ("server_finalidade", "server_finalidade"), ("server_mysql", "server_mysql"),
        ("server_status", "server_status"), ("server_proc", "server_proc"),
        ("server_conteudo", "server_conteudo"), ("server_familia", "server_familia"),
    ]),
    ("list_resp_crono", "list_resp_crono", [
        ("resp_id", "resp_id"), ("resp_nome", "resp_nome"),
    ]),
    ("list_tip_resp", "list_tip_resp", [("tip_resp", "tip_resp")]),
    ("list_url_status", "list_url_status", [("url_status", "url_status")]),
    # "usuarios" nao entra aqui: a aba tem granularidade de (pessoa, tabela que
    # ela pode editar), normalizada em duas tabelas -- ver load_usuarios() abaixo.
    ("indices_economicos", "index", [
        ("index_cod", "index_cod"), ("index_nome", "index_nome"), ("index_ano", "index_ano"),
        ("index_mes", "index_mes"), ("index_vlr", "index_vlr"),
    ]),
    ("fornecedores", "fornecedores", [
        ("fornecedor_id", "fornecedor_id"), ("fornecedor_area", "fornecedor_area"),
        ("fornecedor_nome", "fornecedor_nome"), ("fornecedor_cnpj", "fornecedor_cnpj"),
    ]),
    ("urls", "urls", [   # entra logo apos clientes/produtos/servidores -> dispara os triggers de cliente_status
        ("url_id", "url_id"), ("cliente_id", "cliente_id"), ("url_path", "url_path"),
        ("url_server", "server_id"), ("url_prod", "produto_id"),
        ("url_status", "url_status"), ("url_dt_status", "url_dt_status"), ("url_exc", "url_exc"),
        ("url_dt_exc", "url_dt_exc"), ("url_pasta_raiz", "url_pasta_raiz"),
        ("url_pasta_anexos", "url_pasta_anexos"), ("urb_bd", "urb_bd"), ("url_obs", "url_obs"),
    ]),
    ("pessoas", "pessoas", [   # hierarquia auto-referenciada -> FK fica OFF durante a carga toda
        ("pessoa_id", "pessoa_id"), ("pessoa_nome", "pessoa_nome"), ("pessoa_status", "pessoa_status"),
        ("pessoa_funcao", "pessoa_funcao"), ("pessoa_mail", "pessoa_mail"), ("pessoa_fone", "pessoa_fone"),
        ("pessoa_diretor", "pessoa_diretor"), ("pessoa_ger_exec", "pessoa_ger_exec"),
        ("pessoa_ger", "pessoa_ger"), ("pessoa_lider", "pessoa_lider"), ("pessoa_squad", "pessoa_squad"),
        ("pessoa_billable", "pessoa_billable"),
    ]),
    ("ferias_marcacao", "ferias_marcacao", [
        ("feriasm_id", "feriasm_id"), ("pessoa_id", "pessoa_id"), ("feriasm_tipo", "feriasm_tipo"),
        ("feriasm_ini", "feriasm_ini"), ("feriasm_fim", "feriasm_fim"),
    ]),
    ("contatos", "contatos", [
        ("contato_id", "contato_id"), ("cliente_id", "cliente_id"), ("contato_nome", "contato_nome"),
        ("contato_mail", "contato_mail"), ("contato_fone", "contato_fone"), ("contato_status", "contato_status"),
    ]),
    ("cart_mes", "cart_mes", [
        ("cart_mes_id", "cart_mes_id"), ("cart_ano_mes", "cart_ano_mes"),
        ("cart_vigencia_ativa", "cart_vigencia_ativa"),
    ]),
    ("precos_cliente", "precos_cliente", [
        ("pc_id", "pc_id"), ("cliente_id", "cliente_id"), ("produto_id", "produto_id"),
        ("cart_mes_id", "cart_mes_id"), ("pc_dat_niver", "pc_dat_niver"),
        ("pc_dat_ult_reajuste", "pc_dat_ult_reajuste"),  # AVISO se a planilha-modelo nao tiver -- ok, coluna nova
        ("pc_cod_index", "pc_cod_index"),
        # pc_tip_vlr não entra mais aqui -- moveu pra clientes.cliente_tip_vlr (backfill_cliente_tip_vlr)
        ("pc_vlr_franquia", "pc_vlr_franquia"), ("pc_vlr_unit", "pc_vlr_unit"),
        ("pc_fx1_lim", "pc_fx1_lim"), ("pc_fx2_lim", "pc_fx2_lim"), ("pc_fx3_lim", "pc_fx3_lim"),
        ("pc_fx4_lim", "pc_fx4_lim"), ("pc_fx5_lim", "pc_fx5_lim"), ("pc_fx1_vlr", "pc_fx1_vlr"),
        ("pc_fx2_vlr", "pc_fx2_vlr"), ("pc_fx3_vlr", "pc_fx3_vlr"), ("pc_fx4_vlr", "pc_fx4_vlr"),
        ("pc_fx5_vlr", "pc_fx5_vlr"),
    ]),
    ("consumo_ana", "consumo_ana", [
        ("consumo_id", "consumo_id"), ("cliente_id", "cliente_id"), ("produto_id", "produto_id"),
        ("cart_mes_id", "cart_mes_id"), ("consumo_data", "consumo_data"), ("consumo_qtd", "consumo_qtd"),
        ("consumo_det", "consumo_det"), ("consumo_consit", "consumo_consit"),
    ]),
    ("faturamento", "faturamento", [
        ("fat_id", "fat_id"), ("cart_mes_id", "cart_mes_id"), ("cliente_id", "cliente_id"),
        ("fat_dat_venc", "fat_dat_venc"), ("fat_cod_venc_protheus", "fat_cod_venc_protheus"),
        ("fat_num_nfe", "fat_num_nfe"), ("fat_num_rps", "fat_num_rps"), ("fat_obs", "fat_obs"),
    ]),
    ("carteira", "carteira", [
        ("cart_id", "cart_id"), ("cliente_id", "cliente_id"), ("cart_mes_id", "cart_mes_id"),
        ("cart_qtd", "cart_qtd"), ("cart_vlr", "cart_vlr"), ("cart_pdd", "cart_pdd"),
        ("cart_sem_pdd", "cart_sem_pdd"), ("cart_fat", "cart_fat"), ("cart_qtd_mes", "cart_qtd_mes"),
        ("cart_emprestimos_mes", "cart_emprestimos_mes"), ("cart_ult_def", "cart_ult_def"),
        ("cart_data_base", "cart_data_base"), ("cart_dat_extracao", "cart_dat_extracao"),
        ("cart_rds", "cart_rds"), ("cart_db", "cart_db"), ("cart_prod", "cart_prod"),
        ("cart_url_plan_analitica", "cart_url_plan_analitica"),
    ]),
    ("resp", "resp", [
        ("resp_id", "resp_id"), ("cliente_id", "cliente_id"), ("resp_tipo", "resp_tipo"),
        ("pessoa_id", "pessoa_id"),
    ]),
    ("escala", "escala", [
        ("escala_id", "escala_id"), ("pessoa_id", "pessoa_id"), ("escala_data", "escala_data"),
        ("escala_hora_ini", "escala_hora_ini"), ("escala_hora_fim", "escala_hora_fim"),
    ]),
    ("portfolios", "port", [
        ("port_id", "port_id"), ("cliente_id", "cliente_id"), ("port_tipo", "port_tipo"),
        ("port_nome", "port_nome"), ("port_pm", "port_pm"), ("port_diretorio", "port_diretorio"),
        ("port_status", "port_status"), ("port_pdf", "port_pdf"),
    ]),
    ("crono", "crono", [
        ("crono_id", "crono_id"), ("port_id", "port_id"), ("crono_grupo", "crono_grupo"),
        ("crono_topico", "crono_topico"), ("crono_tipo", "crono_tipo"), ("crono_atividade", "crono_atividade"),
        ("crono_inicio", "crono_inicio"), ("crono_fim", "crono_fim"), ("crono_replan", "crono_replan"),
        ("crono_esforço", "crono_esforco"), ("crono_perc_atual", "crono_perc_atual"),
        ("crono_hh_orc", "crono_hh_orc"), ("crono_hh_real", "crono_hh_real"), ("crono_status", "crono_status"),
        ("resp_id", "resp_id"), ("crono_demanda_1", "crono_demanda_1"), ("crono_demanda_2", "crono_demanda_2"),
        ("crono_demanda_3", "crono_demanda_3"), ("crono_relat", "crono_relat"),
    ]),
    ("propostas", "propostas", [
        ("proposta_id", "proposta_id"), ("cliente_id", "cliente_id"), ("proposta_chamado", "proposta_chamado"),
        ("proposta_demanda", "proposta_demanda"), ("proposta_nome", "proposta_nome"),
        ("proposta_desc", "proposta_desc"), ("proposta_hh", "proposta_hh"), ("proposta_vlr", "proposta_vlr"),
        ("proposta_status", "proposta_status"), ("proposta_anexo", "proposta_anexo"),
    ]),
    ("forn_contratos", "forn_contratos", [
        ("forn_cont_id", "forn_cont_id"), ("fornecedor_id", "fornecedor_id"),
        ("forn_cont_num_contrato", "forn_cont_num_contrato"), ("forn_cont_tipo", "forn_cont_tipo"),
        ("forn_cont_nivel", "forn_cont_nivel"), ("forn_cont_aloc", "forn_cont_aloc"),
        ("forn_cont_qtd_prf", "forn_cont_qtd_prf"), ("forn_cont_desc", "forn_cont_desc"),
        ("forn_cont_tip_vlr", "forn_cont_tip_vlr"), ("forn_cont_vlr_mes", "forn_cont_vlr_mes"),
        ("forn_cont_dt_ini", "forn_cont_dt_ini"), ("forn_cont_dt_fim", "forn_cont_dt_fim"),
        ("forn_cont_ind_reaj", "forn_cont_ind_reaj"), ("forn_cont_status", "forn_cont_status"),
        ("pessoa_id", "pessoa_id"),
    ]),
    ("forn_pagadoria", "forn_pagadoria", [
        ("forn_pag_id", "forn_pag_id"), ("fornecedor_id", "fornecedor_id"), ("forn_pag_resp", "forn_pag_resp"),
        ("forn_pag_tipo", "forn_pag_tipo"), ("forn_pag_tipo_detalhado", "forn_pag_tipo_detalhado"),
        ("forn_pag_competencia", "forn_pag_competencia"), ("forn_pag_dat", "forn_pag_dat"),
        ("forn_pag_nome_prf", "forn_pag_nome_prf"), ("forn_pag_qtd", "forn_pag_qtd"),
        ("forn_pag_vlr_unit", "forn_pag_vlr_unit"), ("forn_pag_tot_bruto", "forn_pag_tot_bruto"),
        ("forn_pag_tot_liq", "forn_pag_tot_liq"),
        ("forn_pag_vlr_pag_cliente_bruto", "forn_pag_vlr_pag_cliente_bruto"),
        ("forn_pag_vlr_pag_cliente_liq", "forn_pag_vlr_pag_cliente_liq"),
        ("forn_pag_vlr_receita_bruta", "forn_pag_vlr_receita_bruta"),
        ("forn_pag_vlr_receita_liq", "forn_pag_vlr_receita_liq"), ("forn_pag_obs", "forn_pag_obs"),
    ]),
    ("anexos", "anexos", [
        ("anexo_id", "anexo_id"), ("cliente_id", "cliente_id"), ("fornecedor_id", "fornecedor_id"),
        ("anexo_nome", "anexo_nome"), ("anexo_data", "anexo_data"), ("anexo_arquivo", "anexo_arquivo"),
    ]),
]


def load_usuarios(wb, conn):
    """usuarios: 1 linha por (pessoa, tabela liberada) na planilha, mas o modelo de
    permissão por tabela foi substituído por permissão por MENU do sistema (ver
    usuarios_permissoes_menu no schema.sql) -- não existe mapeamento direto entre
    user_tabela (coluna antiga da planilha) e menu_key, então essa massa de teste só
    carrega usuarios (1 linha por pessoa, user_mail unico); as permissões por menu
    ficam pra configurar manualmente no Admin."""
    ws = wb["usuarios"]
    header = [c.value for c in ws[1]]
    idx = {name: i for i, name in enumerate(header)}

    usuarios_rows = {}  # user_mail -> (user_id, user_nome, user_mail, user_status)

    for row in ws.iter_rows(min_row=2, values_only=True):
        user_mail = clean(row[idx["user_mail"]])
        if user_mail is None and all(v is None for v in row):
            continue
        user_nome = clean(row[idx["user_nome"]])
        user_status = clean(row[idx["user_status"]])

        if user_mail not in usuarios_rows:
            usuarios_rows[user_mail] = (len(usuarios_rows) + 1, user_nome, user_mail, user_status)

    conn.executemany(
        "INSERT INTO usuarios (user_id, user_nome, user_mail, user_status) VALUES (?, ?, ?, ?)",
        list(usuarios_rows.values()),
    )
    conn.commit()
    print(f"  {'usuarios':25s} <- {'usuarios':20s} {len(usuarios_rows):>6d} linhas")
    return len(usuarios_rows)


def backfill_cliente_tip_vlr(wb, conn):
    """cliente_tip_vlr não existe na aba "clientes" -- na planilha original o regime
    (BRUTO/LIQUIDO) era parametrizado por linha de precos_cliente (pc_tip_vlr), decisão
    revista nesta sessão pra evitar cliente com produtos em regimes diferentes. Backfill:
    lê pc_tip_vlr direto da aba "precos_cliente" (não vai mais pra tabela precos_cliente)
    e usa o primeiro valor não vazio encontrado por cliente. Se o mesmo cliente tiver
    linhas com regimes diferentes na planilha, fica registrado o aviso abaixo -- não
    há como saber qual delas é a correta sem confirmar com o usuário."""
    ws = wb["precos_cliente"]
    header = [c.value for c in ws[1]]
    idx = {name: i for i, name in enumerate(header)}
    if "pc_tip_vlr" not in idx or "cliente_id" not in idx:
        return

    tip_vlr_by_cliente = {}
    conflitos = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        cliente_id = clean(row[idx["cliente_id"]])
        tip_vlr = clean(row[idx["pc_tip_vlr"]])
        if cliente_id is None or tip_vlr is None:
            continue
        existente = tip_vlr_by_cliente.get(cliente_id)
        if existente is None:
            tip_vlr_by_cliente[cliente_id] = tip_vlr
        elif existente != tip_vlr:
            conflitos.add(cliente_id)

    conn.executemany(
        "UPDATE clientes SET cliente_tip_vlr = ? WHERE cliente_id = ?",
        [(tip_vlr, cliente_id) for cliente_id, tip_vlr in tip_vlr_by_cliente.items()],
    )
    conn.commit()
    print(f"  {'clientes.cliente_tip_vlr':25s} <- {'precos_cliente':20s} {len(tip_vlr_by_cliente):>6d} clientes")
    if conflitos:
        print(f"  [AVISO] {len(conflitos)} cliente(s) com pc_tip_vlr divergente entre linhas de preço "
              f"(ficou o primeiro valor encontrado): {sorted(conflitos)}")


def main():
    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    if os.path.exists(DB_PATH):
        os.remove(DB_PATH)

    conn = sqlite3.connect(DB_PATH)

    for fname in ["schema.sql", "views.sql", "triggers.sql"]:
        with open(os.path.join(SQL_DIR, fname), encoding="utf-8") as f:
            conn.executescript(f.read())
    print("schema + views + triggers aplicados")

    # schema.sql liga "PRAGMA foreign_keys = ON" -- desligamos de novo pra carga,
    # porque pessoas tem hierarquia auto-referenciada (referencias pra frente).
    conn.execute("PRAGMA foreign_keys = OFF")

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

    total_rows = 0
    for table_name, sheet_name, col_map in TABLES:
        ws = wb[sheet_name]
        header = [c.value for c in ws[1]]
        header_idx = {name: i for i, name in enumerate(header)}

        missing = [src for src, _ in col_map if src not in header_idx]
        if missing:
            print(f"  [AVISO] {sheet_name}: colunas não encontradas no cabeçalho: {missing}")

        usable_map = [(src, dst) for src, dst in col_map if src in header_idx]
        src_indexes = [header_idx[src] for src, _ in usable_map]
        dst_cols = [dst for _, dst in usable_map]

        not_null_cols = {
            r[1] for r in conn.execute(f"PRAGMA table_info({table_name})").fetchall() if r[3]
        }
        not_null_positions = [i for i, dst in enumerate(dst_cols) if dst in not_null_cols]

        placeholders = ", ".join(["?"] * len(dst_cols))
        col_list = ", ".join(dst_cols)
        insert_sql = f"INSERT INTO {table_name} ({col_list}) VALUES ({placeholders})"

        rows_to_insert = []
        skipped = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(row[i] is None for i in src_indexes):
                continue  # linha totalmente vazia
            values = [clean(row[i]) for i in src_indexes]
            if any(values[i] is None for i in not_null_positions):
                skipped += 1  # linha lixo/placeholder: NOT NULL da coluna ficou vazio apos limpeza
                continue
            rows_to_insert.append(values)

        conn.executemany(insert_sql, rows_to_insert)
        conn.commit()
        print(f"  {table_name:25s} <- {sheet_name:20s} {len(rows_to_insert):>6d} linhas")
        if skipped:
            print(f"  [AVISO] {table_name}: {skipped} linha(s) ignorada(s) (coluna NOT NULL vazia apos limpeza)")
        total_rows += len(rows_to_insert)

        if table_name == "list_url_status":
            total_rows += load_usuarios(wb, conn)

        if table_name == "precos_cliente":
            backfill_cliente_tip_vlr(wb, conn)

    conn.execute("PRAGMA foreign_keys = ON")
    conn.commit()
    conn.close()
    print(f"\nTotal importado: {total_rows} linhas em {len(TABLES)} tabelas")


if __name__ == "__main__":
    main()
